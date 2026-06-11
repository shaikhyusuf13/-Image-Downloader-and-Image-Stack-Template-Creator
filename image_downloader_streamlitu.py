"""
Image Downloader & Image Stack Template Creator — Streamlit Edition
Mirrors all functionality of the tkinter v5 GUI:
  - Excel column mapping (auto-detect + manual override)
  - Three naming conventions
  - Force JPG conversion (requires Pillow)
  - Parallel downloads with progress tracking
  - Image Stack Import Template generation
  - Failure report (.xlsx) generation
  - In-browser ZIP download of all output files
"""

import io
import os
import re
import threading
import zipfile
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from tempfile import mkdtemp
from urllib.parse import urlparse

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Constants ──────────────────────────────────────────────────────────────────
APP_TITLE = "Image Downloader & Image Stack Template Creator"
DEFAULT_SHEET = 0
TIMEOUT = 30
MAX_WORKERS = 12

IMPORT_TYPES = ["Create/Edit", "Edit", "Remove Image", "Delete"]
TEMPLATE_HEADERS = ["Import Type", "Collection Folder", "Image Stack Group", "Filename", "Image Stack Order"]
COLLECTION_FOLDER_SOURCES = ["Master ID", "MPN", "Custom value", "Leave blank"]
LAST_IMAGE_END_OF_ROW = "(end of row)"

NAMING_OPTIONS = {
    "1": "MasterID _ MPN _ Marketplace _ Country _ ISP01",
    "3": "MPN _ MasterID _ Marketplace _ Country _ ISP01",
    "2": "MPN _ MasterID _ Marketplace _ Country _ ISP_01",
}


# ── Pure helpers (shared with tkinter version) ─────────────────────────────────

def sanitize_filename(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip(" ._")


def safe_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def normalize_id(value) -> str:
    text = safe_text(value)
    if not text:
        return ""
    try:
        num = float(text)
        if num.is_integer():
            return str(int(num))
    except Exception:
        pass
    return text


def get_extension_from_response(url: str, response: requests.Response) -> str:
    content_type = (response.headers.get("Content-Type") or "").lower()
    content_map = {
        "image/jpeg": ".jpg", "image/jpg": ".jpg", "image/png": ".png",
        "image/webp": ".webp", "image/gif": ".gif", "image/bmp": ".bmp",
        "image/tiff": ".tif", "image/svg+xml": ".svg",
    }
    if content_type in content_map:
        return content_map[content_type]
    parsed = urlparse(url)
    suffix = Path(parsed.path).suffix.lower()
    if suffix in [".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".svg"]:
        if suffix == ".jpeg":
            return ".jpg"
        if suffix == ".tiff":
            return ".tif"
        return suffix
    return ".jpg"


def build_filename(convention, country, master_id, mpn, image_index, marketplace=""):
    serial = f"ISP{image_index:02d}"
    if convention == "2":
        serial_spaced = f"ISP_{image_index:02d}"
        return sanitize_filename(f"{mpn}_{master_id}_{marketplace}_{country}_{serial_spaced}")
    if convention == "3":
        return sanitize_filename(f"{mpn}_{master_id}_{marketplace}_{country}_{serial}")
    return sanitize_filename(f"{master_id}_{mpn}_{marketplace}_{country}_{serial}")


def parse_image_stack_order(filename: str, fallback=None):
    if not filename:
        return fallback
    base = filename.rsplit(".", 1)[0] if "." in filename else filename
    m = re.search(r"ISP_?(\d+)", base, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return fallback


# ── URL validation ─────────────────────────────────────────────────────────────

PLACEHOLDER_TOKENS = {"", "-", "—", "n/a", "na", "none", "null", "missing", "tbd", "?", "."}


def is_blank_or_placeholder(text: str) -> bool:
    if text is None:
        return True
    return str(text).strip().lower() in PLACEHOLDER_TOKENS


def validate_url(url: str):
    if is_blank_or_placeholder(url):
        return False, "Missing URL", "Cell is blank or contains a placeholder value"
    u = url.strip()
    if not u.lower().startswith(("http://", "https://")):
        snippet = u[:80] + ("…" if len(u) > 80 else "")
        return False, "Invalid URL", f"URL doesn't begin with http:// or https:// — got: {snippet}"
    if " " in u or "\n" in u or "\t" in u:
        return False, "Invalid URL", "URL contains whitespace characters"
    parsed = urlparse(u)
    if not parsed.netloc:
        return False, "Invalid URL", "URL has no host"
    return True, None, None


def categorize_download_error(exc):
    try:
        from PIL import UnidentifiedImageError
        if isinstance(exc, UnidentifiedImageError):
            return "Invalid Image", "Server returned bytes that don't form a valid image"
    except ImportError:
        pass
    if isinstance(exc, requests.exceptions.Timeout):
        return "Timeout", f"Server didn't respond within {TIMEOUT}s"
    if isinstance(exc, requests.exceptions.SSLError):
        return "SSL Error", str(exc).split("\n", 1)[0]
    if isinstance(exc, requests.exceptions.TooManyRedirects):
        return "Too Many Redirects", str(exc)
    if isinstance(exc, requests.exceptions.HTTPError):
        resp = getattr(exc, "response", None)
        if resp is not None:
            return f"HTTP {resp.status_code}", f"{resp.status_code} {resp.reason}"
        return "HTTP Error", str(exc)
    if isinstance(exc, requests.exceptions.ConnectionError):
        msg = str(exc)
        if "Name or service not known" in msg or "getaddrinfo failed" in msg:
            return "Host Not Found", "Domain could not be resolved"
        if "Connection refused" in msg:
            return "Connection Refused", "Server refused the connection"
        return "Connection Error", msg.split("\n", 1)[0][:200]
    if isinstance(exc, (requests.exceptions.MissingSchema, requests.exceptions.InvalidURL,
                        requests.exceptions.InvalidSchema)):
        return "Invalid URL", str(exc)
    if isinstance(exc, RuntimeError) and "Pillow" in str(exc):
        return "Conversion Error", str(exc)
    return "Other Error", str(exc).split("\n", 1)[0][:200] or exc.__class__.__name__


# ── JPG conversion ─────────────────────────────────────────────────────────────

def response_to_jpg_bytes(response, quality: int = 92) -> bytes:
    """Convert any image response to JPG bytes in memory — no disk I/O."""
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("Pillow is required for 'Force JPG'. Install: pip install Pillow") from exc
    buf = io.BytesIO()
    for chunk in response.iter_content(chunk_size=8192):
        if chunk:
            buf.write(chunk)
    buf.seek(0)
    img = Image.open(buf)
    img.load()
    if img.mode in ("RGBA", "LA"):
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1])
        img = bg
    elif img.mode == "P":
        img = img.convert("RGBA")
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1])
        img = bg
    elif img.mode != "RGB":
        img = img.convert("RGB")
    out = io.BytesIO()
    img.save(out, format="JPEG", quality=quality, optimize=True)
    return out.getvalue()


# ── Excel helpers ──────────────────────────────────────────────────────────────

_FIELD_PATTERNS = {
    "master_id":   [r"master\s*[-_ ]?\s*id", r"^id$", r"product\s*id"],
    "mpn":         [r"^mpn$", r"manufacturer.*part", r"part\s*number", r"part\s*no"],
    "image_count": [r"image\s*count", r"img\s*count", r"images?\s*total", r"num.*image",
                    r"^count$", r"no\.?\s*of\s*image"],
    "first_image": [r"image\s*1\b", r"img\s*1\b", r"image[-_ ]?url[-_ ]?1\b",
                    r"image\s*one\b", r"^image$", r"^url$"],
}


def read_excel_columns_from_bytes(file_bytes, sheet=None):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet] if sheet < len(wb.worksheets) else wb.active
        else:
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        columns = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for idx, value in enumerate(row):
                letter = get_column_letter(idx + 1)
                name = "" if value is None else str(value).strip()
                short = (name[:35] + "…") if len(name) > 36 else name
                label = f"{letter} — {short}" if short else f"{letter} — (unnamed)"
                columns.append({"letter": letter, "index": idx, "name": name, "label": label})
            break
        return columns
    finally:
        wb.close()


def auto_detect_mappings(columns):
    found = {}
    for field, patterns in _FIELD_PATTERNS.items():
        for col in columns:
            haystack = col["name"].lower()
            if not haystack:
                continue
            for pat in patterns:
                if re.search(pat, haystack, flags=re.IGNORECASE):
                    found[field] = col
                    break
            if field in found:
                break
    positional_defaults = {"master_id": 0, "mpn": 1, "image_count": 2, "first_image": 3}
    for field, default_idx in positional_defaults.items():
        if field not in found and default_idx < len(columns):
            found[field] = columns[default_idx]
    return found


def resolve_column_index(label_value, excel_columns, default_index=None):
    value = (label_value or "").strip()
    if not value or value == LAST_IMAGE_END_OF_ROW:
        return None
    for col in excel_columns:
        if col["label"] == value:
            return col["index"]
    m = re.match(r"^\s*([A-Za-z]+)\b", value)
    if m:
        from openpyxl.utils import column_index_from_string
        try:
            return column_index_from_string(m.group(1).upper()) - 1
        except Exception:
            pass
    return default_index


# ── Workbook writers ───────────────────────────────────────────────────────────

def generate_image_stack_workbook_bytes(records, import_type, stack_group):
    wb = Workbook()
    tmpl = wb.active
    tmpl.title = "Image Stack Import Template"
    border = Border(
        left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))
    tmpl["A1"] = "Required"
    tmpl["A1"].fill = PatternFill("solid", fgColor="F4CCCC")
    tmpl["A1"].font = Font(bold=True)
    tmpl["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = tmpl.cell(row=2, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="FCE5CD")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    start_row = 3
    for i, rec in enumerate(records):
        r = start_row + i
        tmpl.cell(row=r, column=1, value=rec.get("import_type", import_type))
        tmpl.cell(row=r, column=2, value=rec.get("collection_folder", ""))
        tmpl.cell(row=r, column=3, value=rec.get("stack_group", stack_group))
        tmpl.cell(row=r, column=4, value=rec.get("filename", ""))
        tmpl.cell(row=r, column=5, value=rec.get("image_stack_order", ""))
        for c in range(1, 6):
            tmpl.cell(row=r, column=c).alignment = Alignment(vertical="center")
    last_row = max(start_row + len(records) - 1, 8)
    dv_formula = '"' + ",".join(IMPORT_TYPES) + '"'
    dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)
    dv.add(f"A{start_row}:A{last_row + 5}")
    tmpl.add_data_validation(dv)
    widths_tmpl = {"A": 34, "B": 28, "C": 41, "D": 35, "E": 33}
    for col, w in widths_tmpl.items():
        tmpl.column_dimensions[col].width = w
    tmpl.row_dimensions[1].height = 20
    tmpl.row_dimensions[2].height = 22
    tmpl.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_failure_report_bytes(issues):
    wb = Workbook()
    ws = wb.active
    ws.title = "Download Issues"
    headers = ["Row", "Master ID", "MPN", "Image #", "URL", "Phase", "Issue Category", "Detail"]
    header_fill = PatternFill("solid", fgColor="FCE5CD")
    border = Border(
        left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    phase_fills = {
        "preflight": PatternFill("solid", fgColor="FCE4E4"),
        "download": PatternFill("solid", fgColor="FFF2CC"),
    }
    for r_idx, item in enumerate(issues, start=2):
        ws.cell(row=r_idx, column=1, value=item.get("row"))
        ws.cell(row=r_idx, column=2, value=item.get("master_id"))
        ws.cell(row=r_idx, column=3, value=item.get("mpn"))
        ws.cell(row=r_idx, column=4, value=item.get("image_index"))
        ws.cell(row=r_idx, column=5, value=item.get("url"))
        ws.cell(row=r_idx, column=6, value=item.get("phase", ""))
        ws.cell(row=r_idx, column=7, value=item.get("category", ""))
        ws.cell(row=r_idx, column=8, value=item.get("detail", ""))
        fill = phase_fills.get(item.get("phase", ""))
        for col in range(1, 9):
            cell = ws.cell(row=r_idx, column=col)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
    widths = {"A": 7, "B": 18, "C": 18, "D": 9, "E": 55, "F": 11, "G": 22, "H": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    last_row = max(len(issues) + 1, 1)
    if last_row > 1:
        ws.auto_filter.ref = f"A1:H{last_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Core download engine (returns results, no UI coupling) ─────────────────────

def run_download_engine(
    file_bytes, sheet_arg,
    country, marketplace, naming, worker_count, force_jpg,
    col_master_id_label, col_mpn_label, col_image_count_label,
    col_first_image_label, col_last_image_label,
    excel_columns,
    progress_callback=None,   # callable(completed, total, message)
    log_callback=None,        # callable(message, level)
):
    def log(msg, level="info"):
        if log_callback:
            log_callback(msg, level)

    results = {
        "tasks_queued": 0,
        "success_count": 0,
        "fail_count": 0,
        "preflight_issues": [],
        "download_failures": [],
        "successful_records": [],
        "downloaded_files": {},   # filename -> bytes
        "error": None,
    }

    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_arg, dtype=object)

        if df.empty:
            results["error"] = "The selected sheet is empty."
            return results

        # Resolve column indices
        master_id_idx = resolve_column_index(col_master_id_label, excel_columns, default_index=0)
        mpn_idx       = resolve_column_index(col_mpn_label,       excel_columns, default_index=1)
        count_idx     = resolve_column_index(col_image_count_label, excel_columns, default_index=2)
        first_img_idx = resolve_column_index(col_first_image_label, excel_columns, default_index=3)
        last_img_idx  = resolve_column_index(col_last_image_label,  excel_columns)

        def _letter(i):
            return get_column_letter(i + 1) if i is not None else "—"

        log(f"Column mapping → Master ID: {_letter(master_id_idx)}  •  "
            f"MPN: {_letter(mpn_idx)}  •  "
            f"Image Count: {_letter(count_idx)}  •  "
            f"First Image: {_letter(first_img_idx)}  •  "
            f"Last Image: {_letter(last_img_idx) if last_img_idx is not None else '(end of row)'}")

        tasks = []
        preflight_issues = results["preflight_issues"]
        skipped_rows = 0

        for row_number, row in df.iterrows():
            excel_row = row_number + 2
            row_len = len(row)

            master_id = normalize_id(row.iloc[master_id_idx]) if master_id_idx < row_len else ""
            mpn = safe_text(row.iloc[mpn_idx]) if mpn_idx < row_len else ""
            image_count_raw = row.iloc[count_idx] if count_idx < row_len else ""

            try:
                declared_count = int(float(image_count_raw)) if safe_text(image_count_raw) else 0
            except Exception:
                declared_count = 0

            end_idx = last_img_idx if last_img_idx is not None else (row_len - 1)
            end_idx = min(end_idx, row_len - 1)
            max_slots = end_idx - first_img_idx + 1

            if max_slots <= 0:
                if declared_count > 0:
                    preflight_issues.append({
                        "row": excel_row, "master_id": master_id, "mpn": mpn,
                        "image_index": "—", "url": "", "phase": "preflight",
                        "category": "Missing Column",
                        "detail": f"Declared {declared_count} images but no image columns exist",
                    })
                continue

            last_filled_pos = 0
            actual_filled = 0
            for col_offset in range(max_slots):
                if safe_text(row.iloc[first_img_idx + col_offset]):
                    last_filled_pos = col_offset + 1
                    actual_filled += 1

            if declared_count == 0 and last_filled_pos == 0:
                skipped_rows += 1
                preflight_issues.append({
                    "row": excel_row, "master_id": master_id, "mpn": mpn,
                    "image_index": "—", "url": "", "phase": "preflight",
                    "category": "No Images",
                    "detail": "Image count is 0 and no URLs found in image columns",
                })
                continue

            effective_max = min(max_slots, max(declared_count, last_filled_pos))

            for image_index in range(1, effective_max + 1):
                col_index = first_img_idx + (image_index - 1)
                if col_index >= row_len:
                    preflight_issues.append({
                        "row": excel_row, "master_id": master_id, "mpn": mpn,
                        "image_index": image_index, "url": "", "phase": "preflight",
                        "category": "Missing Column",
                        "detail": f"Image {image_index} declared but column doesn't exist",
                    })
                    continue

                url = safe_text(row.iloc[col_index])
                valid, cat, detail = validate_url(url)
                if not valid:
                    gap_note = ""
                    if image_index < last_filled_pos:
                        gap_note = f"  (gap — URLs exist at later positions, up to Image {last_filled_pos})"
                    preflight_issues.append({
                        "row": excel_row, "master_id": master_id, "mpn": mpn,
                        "image_index": image_index, "url": url, "phase": "preflight",
                        "category": cat, "detail": detail + gap_note,
                    })
                    continue

                filename_base = build_filename(naming, country, master_id, mpn, image_index, marketplace)
                tasks.append({
                    "row_number": excel_row,
                    "url": url.strip(),
                    "filename_base": filename_base,
                    "master_id": master_id,
                    "mpn": mpn,
                    "image_index": image_index,
                })

        preflight_fail = len([p for p in preflight_issues if p["category"] != "No Images"])
        log(f"Pre-flight complete — {len(tasks)} URLs queued, "
            f"{preflight_fail} URL issues, {skipped_rows} rows skipped.")

        results["tasks_queued"] = len(tasks)
        total = len(tasks)

        if total == 0:
            results["error"] = "no_valid_urls"
            return results

        # ── Download ──
        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=worker_count, pool_maxsize=worker_count)
        session.mount("http://", adapter)
        session.mount("https://", adapter)

        completed = 0
        success_count = 0
        fail_count = 0
        name_collision_counter = {}
        records_lock = threading.Lock()
        download_failures = results["download_failures"]
        successful_records = results["successful_records"]
        downloaded_files = results["downloaded_files"]

        def worker(task):
            url = task["url"]
            response = session.get(url, timeout=TIMEOUT, stream=True)
            response.raise_for_status()

            ext = ".jpg" if force_jpg else get_extension_from_response(url, response)
            file_base = task["filename_base"]
            file_name = file_base + ext

            with records_lock:
                if file_name in downloaded_files:
                    count = name_collision_counter.get(file_name, 1)
                    while True:
                        alt_name = f"{file_base}_{count}{ext}"
                        if alt_name not in downloaded_files:
                            file_name = alt_name
                            name_collision_counter[task["filename_base"] + ext] = count + 1
                            break
                        count += 1

            if force_jpg:
                file_bytes_out = response_to_jpg_bytes(response)
            else:
                buf = io.BytesIO()
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        buf.write(chunk)
                file_bytes_out = buf.getvalue()

            return file_name, file_bytes_out

        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            future_map = {executor.submit(worker, task): task for task in tasks}
            for future in as_completed(future_map):
                task = future_map[future]
                completed += 1
                if progress_callback:
                    progress_callback(completed, total, f"Downloading… {completed}/{total}")
                try:
                    file_name, file_bytes_out = future.result()
                    success_count += 1
                    with records_lock:
                        downloaded_files[file_name] = file_bytes_out
                        successful_records.append({
                            "master_id": task["master_id"],
                            "mpn": task["mpn"],
                            "image_index": task["image_index"],
                            "row_number": task["row_number"],
                            "filename": file_name,
                        })
                    log(f"✔  Row {task['row_number']:>4} | Image {task['image_index']:>2} | {file_name}", "ok")
                except Exception as exc:
                    fail_count += 1
                    category, detail = categorize_download_error(exc)
                    download_failures.append({
                        "row": task["row_number"],
                        "master_id": task["master_id"],
                        "mpn": task["mpn"],
                        "image_index": task["image_index"],
                        "url": task["url"],
                        "phase": "download",
                        "category": category,
                        "detail": detail,
                    })
                    log(f"✘  {category:<18} Row {task['row_number']:>4} | Image {task['image_index']:>2} | {detail}", "fail")

        results["success_count"] = success_count
        results["fail_count"] = fail_count

    except Exception as exc:
        results["error"] = str(exc)

    return results


# ── Streamlit UI ───────────────────────────────────────────────────────────────

def collection_value_for(source, master_id, mpn, custom_val):
    if source == "Master ID":
        return master_id
    if source == "MPN":
        return mpn
    if source == "Custom value":
        return custom_val.strip()
    return ""


def main():
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="📦",
        layout="wide",
    )

    # ── Custom CSS ──
    st.markdown("""
    <style>
        .block-container { padding-top: 2rem; padding-bottom: 2rem; }
        .stRadio > div { gap: 0.4rem; }
        div[data-testid="stExpander"] { border: 1px solid #dfe6ee; border-radius: 8px; }
        .log-box {
            background: #0f172a; color: #e2e8f0;
            font-family: 'Cascadia Mono', 'Consolas', monospace;
            font-size: 12px; padding: 12px 16px;
            border-radius: 8px; max-height: 340px;
            overflow-y: auto; white-space: pre-wrap; word-break: break-all;
        }
        .log-ok   { color: #4ade80; }
        .log-fail { color: #f87171; }
        .log-info { color: #93c5fd; }
        .log-muted{ color: #94a3b8; }
    </style>
    """, unsafe_allow_html=True)

    st.title("📦 Image Downloader & Image Stack Template Creator")
    st.caption("Upload your Excel file, configure settings, download images, and auto-build the Image Stack template.")

    # ── Session state init ──
    if "excel_columns" not in st.session_state:
        st.session_state.excel_columns = []
    if "log_lines" not in st.session_state:
        st.session_state.log_lines = []
    if "run_results" not in st.session_state:
        st.session_state.run_results = None

    # ═══════════════════════════════════════════════════════════════════════════
    # Row 1: Files + Naming Convention
    # ═══════════════════════════════════════════════════════════════════════════
    col_files, col_naming = st.columns(2, gap="large")

    with col_files:
        st.subheader("📁 Files")
        uploaded_file = st.file_uploader(
            "Excel file (.xlsx / .xls / .xlsm)",
            type=["xlsx", "xls", "xlsm"],
            help="The spreadsheet containing Master IDs, MPNs, and image URLs.",
        )
        sheet_input = st.text_input(
            "Sheet name / index",
            value="0",
            help="Sheet index (0 = first sheet) or sheet name.",
        )

    with col_naming:
        st.subheader("🏷️ Naming Convention")
        naming_choice = st.radio(
            "Choose filename format",
            options=list(NAMING_OPTIONS.keys()),
            format_func=lambda k: NAMING_OPTIONS[k],
            label_visibility="collapsed",
        )

    # ═══════════════════════════════════════════════════════════════════════════
    # Column Mapping (full width, shown after file upload)
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("📋 Excel Column Mapping")

    file_bytes = None
    sheet_arg = DEFAULT_SHEET

    if uploaded_file:
        file_bytes = uploaded_file.read()
        try:
            sheet_arg = int(sheet_input.strip()) if sheet_input.strip() else DEFAULT_SHEET
        except ValueError:
            sheet_arg = sheet_input.strip()

        try:
            columns = read_excel_columns_from_bytes(file_bytes, sheet=sheet_arg)
            st.session_state.excel_columns = columns
        except Exception as e:
            st.warning(f"Couldn't read column headers: {e}. Positional defaults will be used.")
            st.session_state.excel_columns = []

    excel_columns = st.session_state.excel_columns

    if excel_columns:
        detected = auto_detect_mappings(excel_columns)
        col_labels = [c["label"] for c in excel_columns]
        last_img_options = [LAST_IMAGE_END_OF_ROW] + col_labels

        def default_label(field):
            return detected[field]["label"] if detected.get(field) else col_labels[0]

        st.caption(f"Loaded **{len(excel_columns)} columns** — auto-detected mapping shown below. Override if needed.")

        mc1, mc2, mc3, mc4 = st.columns(4)
        with mc1:
            sel_master_id = st.selectbox("Master ID *", col_labels,
                index=col_labels.index(default_label("master_id")) if default_label("master_id") in col_labels else 0)
        with mc2:
            sel_mpn = st.selectbox("MPN *", col_labels,
                index=col_labels.index(default_label("mpn")) if default_label("mpn") in col_labels else 0)
        with mc3:
            sel_image_count = st.selectbox("Image Count *", col_labels,
                index=col_labels.index(default_label("image_count")) if default_label("image_count") in col_labels else 0)
        with mc4:
            sel_first_image = st.selectbox("First Image *", col_labels,
                index=col_labels.index(default_label("first_image")) if default_label("first_image") in col_labels else 0)

        sel_last_image = st.selectbox(
            "Last Image",
            last_img_options,
            index=0,
            help='Leave as "(end of row)" to include every column from First Image to the end.',
        )
    else:
        st.info("Upload an Excel file above to configure column mapping. Default positional layout will be used: **A=Master ID, B=MPN, C=Image Count, D=Image 1**.")
        sel_master_id = sel_mpn = sel_image_count = sel_first_image = ""
        sel_last_image = LAST_IMAGE_END_OF_ROW

    # ═══════════════════════════════════════════════════════════════════════════
    # Row 2: Product Context + Engine Settings
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    col_ctx, col_eng = st.columns(2, gap="large")

    with col_ctx:
        st.subheader("🌍 Product Context")
        country_name = st.text_input("Country name *", placeholder="e.g. USA")
        marketplace  = st.text_input("Marketplace *",  placeholder="e.g. AMZ")

    with col_eng:
        st.subheader("⚙️ Engine Settings")
        max_workers = st.slider("Download threads", min_value=1, max_value=64, value=MAX_WORKERS)
        force_jpg   = st.checkbox("Force convert all images to JPG",
                                  help="PNG / WebP / GIF / BMP / TIFF re-encoded as JPG. Requires Pillow.")
        save_report = st.checkbox("Generate failure report (.xlsx)", value=True)

    # ═══════════════════════════════════════════════════════════════════════════
    # Image Stack Template settings
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("📚 Image Stack Import Template")

    stack_enabled = st.checkbox("Generate Image Stack template after download", value=True)
    if stack_enabled:
        sc1, sc2, sc3, sc4 = st.columns(4)
        with sc1:
            import_type = st.selectbox("Import Type *", IMPORT_TYPES)
        with sc2:
            stack_group = st.text_input("Image Stack Group *", placeholder="e.g. Main Images")
        with sc3:
            collection_source = st.selectbox("Collection Folder source", COLLECTION_FOLDER_SOURCES)
        with sc4:
            collection_custom = st.text_input(
                "Custom value",
                disabled=(collection_source != "Custom value"),
                placeholder="Only needed for 'Custom value'",
            )
    else:
        import_type = IMPORT_TYPES[0]
        stack_group = ""
        collection_source = COLLECTION_FOLDER_SOURCES[0]
        collection_custom = ""

    # ═══════════════════════════════════════════════════════════════════════════
    # Validation & Run
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")

    def validate():
        errors = []
        if not uploaded_file:
            errors.append("Please upload an Excel file.")
        if not country_name.strip():
            errors.append("Country name is required.")
        if not marketplace.strip():
            errors.append("Marketplace is required.")
        if stack_enabled:
            if not stack_group.strip():
                errors.append("Image Stack Group is required (or untick the template option).")
            if collection_source == "Custom value" and not collection_custom.strip():
                errors.append("Custom Collection Folder value is required.")
        return errors

    run_col, _ = st.columns([1, 3])
    with run_col:
        run_clicked = st.button("🚀 Start Download", type="primary", use_container_width=True)

    if run_clicked:
        errors = validate()
        if errors:
            for e in errors:
                st.error(e)
        else:
            st.session_state.log_lines = []
            st.session_state.run_results = None

            log_lines = []

            def log_cb(msg, level="info"):
                log_lines.append((msg, level))

            progress_bar = st.progress(0, text="Starting…")

            def progress_cb(done, total, msg):
                pct = done / total if total else 0
                progress_bar.progress(pct, text=msg)

            with st.spinner("Running downloads…"):
                results = run_download_engine(
                    file_bytes=file_bytes,
                    sheet_arg=sheet_arg,
                    country=sanitize_filename(country_name.strip()),
                    marketplace=sanitize_filename(marketplace.strip()),
                    naming=naming_choice,
                    worker_count=max_workers,
                    force_jpg=force_jpg,
                    col_master_id_label=sel_master_id,
                    col_mpn_label=sel_mpn,
                    col_image_count_label=sel_image_count,
                    col_first_image_label=sel_first_image,
                    col_last_image_label=sel_last_image,
                    excel_columns=excel_columns,
                    progress_callback=progress_cb,
                    log_callback=log_cb,
                )

            progress_bar.empty()
            st.session_state.log_lines = log_lines
            st.session_state.run_results = results

    # ═══════════════════════════════════════════════════════════════════════════
    # Results
    # ═══════════════════════════════════════════════════════════════════════════
    results = st.session_state.run_results

    if results:
        st.markdown("---")

        if results.get("error") == "no_valid_urls":
            st.error("No valid image URLs found. Check the failure report for details.")
        elif results.get("error"):
            st.error(f"Error: {results['error']}")
        else:
            sc, fc = results["success_count"], results["fail_count"]
            pf = len([p for p in results["preflight_issues"] if p["category"] != "No Images"])

            m1, m2, m3 = st.columns(3)
            m1.metric("✅ Downloaded", sc)
            m2.metric("❌ Failed", fc)
            m3.metric("⚠️ Pre-flight issues", pf)

            # Issue breakdown
            all_issues = results["preflight_issues"] + results["download_failures"]
            if all_issues:
                cat_counts = Counter(it["category"] for it in all_issues if it["category"] != "No Images")
                if cat_counts:
                    with st.expander("Issue breakdown"):
                        for cat, n in sorted(cat_counts.items(), key=lambda kv: -kv[1]):
                            st.write(f"• **{cat}** — {n}")

            # Build ZIP with images + template + report
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                # Images
                for fname, fbytes in results["downloaded_files"].items():
                    zf.writestr(fname, fbytes)

                # Image Stack template
                if stack_enabled and results["successful_records"]:
                    grouped = defaultdict(list)
                    for rec in results["successful_records"]:
                        grouped[(rec["master_id"], rec["mpn"])].append(rec)

                    rows = []
                    ordered_keys = sorted(grouped.keys(),
                                          key=lambda k: min(r["row_number"] for r in grouped[k]))
                    for key in ordered_keys:
                        items = sorted(grouped[key], key=lambda r: r["image_index"])
                        for item in items:
                            rows.append({
                                "import_type": import_type,
                                "collection_folder": collection_value_for(
                                    collection_source, item["master_id"],
                                    item["mpn"], collection_custom),
                                "stack_group": stack_group,
                                "filename": item["filename"],
                                "image_stack_order": parse_image_stack_order(
                                    item["filename"], fallback=item["image_index"]),
                            })

                    template_bytes = generate_image_stack_workbook_bytes(rows, import_type, stack_group)
                    zf.writestr(f"image_stack_template_{timestamp}.xlsx", template_bytes)

                # Failure report
                if save_report and all_issues:
                    report_bytes = write_failure_report_bytes(all_issues)
                    zf.writestr(f"download_report_{timestamp}.xlsx", report_bytes)

            zip_buf.seek(0)

            st.download_button(
                label=f"⬇️ Download all ({sc} images"
                      + (" + template" if stack_enabled and results["successful_records"] else "")
                      + (" + report" if save_report and all_issues else "")
                      + ") as ZIP",
                data=zip_buf.getvalue(),
                file_name=f"images_{timestamp}.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )

            # Individual template download if available
            if stack_enabled and results["successful_records"]:
                grouped = defaultdict(list)
                for rec in results["successful_records"]:
                    grouped[(rec["master_id"], rec["mpn"])].append(rec)
                rows = []
                ordered_keys = sorted(grouped.keys(),
                                      key=lambda k: min(r["row_number"] for r in grouped[k]))
                for key in ordered_keys:
                    items = sorted(grouped[key], key=lambda r: r["image_index"])
                    for item in items:
                        rows.append({
                            "import_type": import_type,
                            "collection_folder": collection_value_for(
                                collection_source, item["master_id"],
                                item["mpn"], collection_custom),
                            "stack_group": stack_group,
                            "filename": item["filename"],
                            "image_stack_order": parse_image_stack_order(
                                item["filename"], fallback=item["image_index"]),
                        })
                tmpl_bytes = generate_image_stack_workbook_bytes(rows, import_type, stack_group)
                st.download_button(
                    label="📄 Download Image Stack Template (.xlsx)",
                    data=tmpl_bytes,
                    file_name=f"image_stack_template_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # Failure report standalone
            if save_report and all_issues:
                report_bytes = write_failure_report_bytes(all_issues)
                st.download_button(
                    label="📄 Download Failure Report (.xlsx)",
                    data=report_bytes,
                    file_name=f"download_report_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ── Activity Log ──────────────────────────────────────────────────────────
    if st.session_state.log_lines:
        st.markdown("---")
        st.subheader("📋 Activity Log")

        level_css = {"ok": "log-ok", "fail": "log-fail", "info": "log-info", "muted": "log-muted"}
        html_lines = []
        for msg, level in st.session_state.log_lines:
            css = level_css.get(level, "")
            escaped = msg.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            if css:
                html_lines.append(f'<span class="{css}">{escaped}</span>')
            else:
                html_lines.append(escaped)

        log_html = '<div class="log-box">' + "<br>".join(html_lines) + "</div>"
        st.markdown(log_html, unsafe_allow_html=True)

    # ── Footer ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        "<div style='text-align:center; color:gray; font-size:13px; padding: 8px 0;'>"
        "Made by <strong>Yusuf Shaikh</strong>"
        "</div>",
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
